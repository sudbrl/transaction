import streamlit as st
import pandas as pd
import dask.dataframe as dd
from openpyxl import load_workbook
from openpyxl.styles import Font
from io import BytesIO
import os
import tempfile

# Correct page config with only valid menu_items keys
st.set_page_config(
    page_title="Transaction Categorizer",
    page_icon="📊",
    initial_sidebar_state="auto",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': None
    }
)

# --- Hide Streamlit UI components ---
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)

# --- Autofit Excel Columns ---
def autofit_excel(writer):
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        for column_cells in worksheet.columns:
            max_length = max((len(str(cell.value)) for cell in column_cells), default=0)
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

# --- Preprocess DataFrame ---
def preprocess_dataframe(df):
    loan_types_to_exclude = [
        'STAFF SOCIAL LOAN', 'STAFF VEHICLE LOAN', 'STAFF HOME LOAN',
        'STAFF FLEXIBLE LOAN', 'STAFF HOME LOAN(COF)', 'STAFF VEHICLE FACILITY LOAN (EVF)'
    ]
    df['Ac Type Desc'] = df['Ac Type Desc'].str.strip().str.upper()
    loan_types_to_exclude = [loan_type.upper() for loan_type in loan_types_to_exclude]
    df = df[~df['Ac Type Desc'].isin(loan_types_to_exclude)]
    df = df[df['Limit'] != 0]
    df = df[~df['Main Code'].isin(['AcType Total', 'Grand Total'])]
    return df

# --- Compare Excel Files ---
def compare_excel_files(df_previous, df_this, writer):
    required_columns = ['Main Code', 'Balance']
    for col in required_columns:
        if col not in df_previous.columns or col not in df_this.columns:
            raise ValueError(f"Missing required column: '{col}'")

    df_previous = preprocess_dataframe(df_previous)
    df_this = preprocess_dataframe(df_this)

    previous_codes = set(df_previous['Main Code'])
    this_codes = set(df_this['Main Code'])

    only_in_previous = df_previous[df_previous['Main Code'].isin(previous_codes - this_codes)]
    only_in_this = df_this[df_this['Main Code'].isin(this_codes - previous_codes)]
    in_both = pd.merge(
        df_previous[['Main Code', 'Balance']],
        df_this[['Main Code', 'Branch Name', 'Name', 'Ac Type Desc', 'Balance']],
        on='Main Code',
        suffixes=('_previous', '_this')
    )
    in_both['Change'] = in_both['Balance_this'] - in_both['Balance_previous']

    opening_sum = df_previous['Balance'].sum()
    settled_sum = only_in_previous['Balance'].sum()
    new_sum = only_in_this['Balance'].sum()
    increase_decrease_sum = in_both['Change'].sum()
    adjusted_sum = opening_sum - settled_sum + new_sum + increase_decrease_sum
    closing_sum = df_this['Balance'].sum()

    reco_data = {
        'Description': ['Opening', 'Settled', 'New', 'Increase/Decrease', 'Adjusted', 'Closing'],
        'Amount': [opening_sum, -settled_sum, new_sum, increase_decrease_sum, adjusted_sum, closing_sum],
        'No of Acs': [len(previous_codes), -len(previous_codes - this_codes), len(this_codes - previous_codes), "", "", len(this_codes)]
    }
    df_reco = pd.DataFrame(reco_data)

    only_in_previous.to_excel(writer, sheet_name='Settled', index=False)
    only_in_this.to_excel(writer, sheet_name='New', index=False)
    in_both[['Main Code', 'Ac Type Desc', 'Branch Name', 'Name', 'Balance_this', 'Balance_previous', 'Change']].to_excel(writer, sheet_name='Movement', index=False)
    df_reco.to_excel(writer, sheet_name='Reco', index=False)

# --- Read Excel Sheets with Dask ---
def read_excel_sheets(file):
    sheets = pd.read_excel(file, sheet_name=None)
    return {sheet_name: dd.from_pandas(sheet_df, npartitions=1) for sheet_name, sheet_df in sheets.items()}

# --- Compare by Ac Type Desc ---
def calculate_common_actype_desc(sheets_1, sheets_2, writer):
    common_actype_present = False
    combined_df = pd.DataFrame()
    for df1 in sheets_1.values():
        for df2 in sheets_2.values():
            if all(col in df1.columns for col in ['Ac Type Desc', 'Balance', 'Main Code', 'Limit']) and \
               all(col in df2.columns for col in ['Ac Type Desc', 'Balance', 'Main Code', 'Limit']):
                common_actype_present = True
                df1 = preprocess_dataframe(df1.compute())
                df2 = preprocess_dataframe(df2.compute())
                df1_grouped = df1.groupby('Ac Type Desc').agg({'Balance': 'sum', 'Ac Type Desc': 'count'})
                df2_grouped = df2.groupby('Ac Type Desc').agg({'Balance': 'sum', 'Ac Type Desc': 'count'})
                df1_grouped.columns = ['Previous Balance Sum', 'Previous Count']
                df2_grouped.columns = ['New Balance Sum', 'New Count']
                combined_df = pd.merge(df1_grouped, df2_grouped, left_index=True, right_index=True, how='outer').fillna(0)
                combined_df['Change'] = combined_df['New Balance Sum'] - combined_df['Previous Balance Sum']
                combined_df['Percent Change'] = ((combined_df['Change'] / combined_df['Previous Balance Sum'].replace({0: pd.NA})) * 100).fillna(0).map('{:.2f}%'.format)
                total_row = pd.DataFrame(combined_df.sum()).transpose()
                total_row.index = ['Total']
                total_row.at['Total', 'Percent Change'] = '{:.2f}%'.format(
                    ((total_row['New Balance Sum'] - total_row['Previous Balance Sum']) / total_row['Previous Balance Sum']).values[0]
                    if total_row['Previous Balance Sum'].values[0] != 0 else 0
                )
                combined_df = pd.concat([combined_df, total_row])

    if common_actype_present:
        combined_df.reset_index().to_excel(writer, sheet_name='Compare', index=False)
        worksheet = writer.sheets['Compare']
        total_row_idx = len(combined_df)
        for col in range(len(combined_df.columns)):
            cell = worksheet.cell(row=total_row_idx + 1, column=col + 1)
            cell.font = Font(bold=True)
            if combined_df.columns[col] == 'Change':
                cell.number_format = '0.00'
    return common_actype_present

# --- Compare by Branch Name ---
def calculate_common_branch_name(sheets_1, sheets_2, writer):
    common_branch_present = False
    combined_df = pd.DataFrame()
    for df1 in sheets_1.values():
        for df2 in sheets_2.values():
            if all(col in df1.columns for col in ['Ac Type Desc', 'Balance', 'Main Code', 'Limit']) and \
               all(col in df2.columns for col in ['Ac Type Desc', 'Balance', 'Main Code', 'Limit']):
                common_branch_present = True
                df1 = preprocess_dataframe(df1.compute())
                df2 = preprocess_dataframe(df2.compute())
                df1_grouped = df1.groupby('Branch Name').agg({'Balance': 'sum', 'Branch Name': 'count'})
                df2_grouped = df2.groupby('Branch Name').agg({'Balance': 'sum', 'Branch Name': 'count'})
                df1_grouped.columns = ['Previous Balance Sum', 'Previous Count']
                df2_grouped.columns = ['New Balance Sum', 'New Count']
                combined_df = pd.merge(df1_grouped, df2_grouped, left_index=True, right_index=True, how='outer').fillna(0)
                combined_df['Change'] = combined_df['New Balance Sum'] - combined_df['Previous Balance Sum']
                combined_df['Percent Change'] = ((combined_df['Change'] / combined_df['Previous Balance Sum'].replace({0: pd.NA})) * 100).fillna(0).map('{:.2f}%'.format)
                total_row = pd.DataFrame(combined_df.sum()).transpose()
                total_row.index = ['Total']
                total_row.at['Total', 'Percent Change'] = '{:.2f}%'.format(
                    ((total_row['New Balance Sum'] - total_row['Previous Balance Sum']) / total_row['Previous Balance Sum']).values[0]
                    if total_row['Previous Balance Sum'].values[0] != 0 else 0
                )
                combined_df = pd.concat([combined_df, total_row])

    if common_branch_present:
        combined_df.reset_index().to_excel(writer, sheet_name='Branch', index=False)
        worksheet = writer.sheets['Branch']
        total_row_idx = len(combined_df)
        for col in range(len(combined_df.columns)):
            cell = worksheet.cell(row=total_row_idx + 1, column=col + 1)
            cell.font = Font(bold=True)
            if combined_df.columns[col] == 'Change':
                cell.number_format = '0.00'
    return common_branch_present

# --- Login Page ---
def login_page():
    st.markdown("""
        <style>
        .login-container {
            max-width: 280px;
            margin: 60px auto;
            padding: 15px 20px;
            background: #f0f2f6;
            border-radius: 6px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.1);
        }
        .login-header {
            font-size: 20px;
            font-weight: 600;
            color: #333;
            margin-bottom: 15px;
            text-align: center;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }
        </style>
        <div class="login-container">
            <div class="login-header">Please Log In</div>
        </div>
    """, unsafe_allow_html=True)

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")

    if submitted:
        if username in st.secrets["auth"] and password == st.secrets["auth"][username]:
            st.session_state["authenticated"] = True
            st.success("Login successful!")
        else:
            st.session_state["authenticated"] = False
            st.error("Invalid username or password.")

# --- Main App UI ---
def app_page():
    st.title("📊 Excel File Comparison Tool")

    st.write("Upload the previous period's Excel file and this period's Excel file to compare them.")
    previous_file = st.file_uploader("Upload Previous Period's Excel File", type=["xlsx"])
    current_file = st.file_uploader("Upload This Period's Excel File", type=["xlsx"])

    if previous_file and current_file:
        st.markdown('<style>div.stButton > button { background-color: #0b0080; color: white; font-weight: bold; }</style>', unsafe_allow_html=True)
        if st.button("Start Processing"):
            with st.spinner("Processing... Please wait."):
                try:
                    with tempfile.NamedTemporaryFile(delete=True, suffix='.xlsx') as tmp_prev, \
                         tempfile.NamedTemporaryFile(delete=True, suffix='.xlsx') as tmp_curr:

                        tmp_prev.write(previous_file.getvalue())
                        tmp_curr.write(current_file.getvalue())

                        previous_wb = load_workbook(tmp_prev.name)
                        current_wb = load_workbook(tmp_curr.name)

                        if len(previous_wb.sheetnames) > 1 or len(current_wb.sheetnames) > 1:
                            st.error("Each workbook should only contain one sheet.")
                        else:
                            df_previous = pd.read_excel(tmp_prev.name)
                            df_this = pd.read_excel(tmp_curr.name)

                            excel_sheets_1 = read_excel_sheets(tmp_prev.name)
                            excel_sheets_2 = read_excel_sheets(tmp_curr.name)

                            output = BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                calculate_common_actype_desc(excel_sheets_1, excel_sheets_2, writer)
                                calculate_common_branch_name(excel_sheets_1, excel_sheets_2, writer)
                                compare_excel_files(df_previous, df_this, writer)
                                autofit_excel(writer)

                            output.seek(0)
                            st.success("Processing completed successfully!")
                            st.download_button(
                                label="Download Comparison Sheet",
                                data=output,
                                file_name="combined_comparison_output.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                except Exception as e:
                    st.error(f"An error occurred during processing: {e}")
                finally:
                    if 'df_previous' in locals(): del df_previous
                    if 'df_this' in locals(): del df_this
                    if 'excel_sheets_1' in locals(): del excel_sheets_1
                    if 'excel_sheets_2' in locals(): del excel_sheets_2
                    if 'output' in locals(): output.close()

# --- Main Entry ---
def main():
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False

    if st.session_state["authenticated"]:
        app_page()
    else:
        login_page()

if __name__ == "__main__":
    main()
